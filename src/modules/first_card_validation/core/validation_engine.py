import os
import re
import traceback
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Import from same package using absolute imports
from .file_parsers import *
from .qr_processor import *
from .excel_generator import *

# Import from utils using absolute path
import sys
import os
sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
from utils.helpers import *

def debug_pcom_content(pcom_path, key_patterns):
    """Debug function to see what's actually in the PCOM file"""
    print(f"\n🔍 DEBUG PCOM FILE: {pcom_path}")
    try:
        with open(pcom_path, 'r', encoding='utf-8', errors='ignore') as file:
            content = file.read()
            lines = content.split('\n')
            
        print("Relevant lines in PCOM file:")
        for i, line in enumerate(lines):
            line_lower = line.lower()
            # Check if line contains any of our target patterns
            if '.define' in line_lower:
                for var in ['%imsi', '%acc', '%puk', '%isc', '%iccid', '%adm', '%home_imsi', '%home_acc', '%dpuk1_card', '%dpuk2_card', '%adm1_card', '%iccid_card']:
                    if var in line_lower:
                        print(f"Line {i}: {line.strip()}")
                        break
                    
    except Exception as e:
        print(f"Error reading PCOM file for debug: {e}")

def extract_from_pcom_enhanced(file_path, patterns):
    """
    Enhanced PCOM extraction - searches entire file for patterns
    """
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
            content = file.read()
        
        # Try multiple patterns
        if isinstance(patterns, str):
            patterns = [patterns]
        
        for pattern in patterns:
            match = re.search(pattern, content, re.IGNORECASE)
            if match:
                print(f"✅ PCOM pattern matched: {pattern} -> {match.group(1)}")
                return match.group(1)
        
        print(f"❌ No PCOM pattern matched from: {patterns}")
        return None
        
    except Exception as e:
        print(f"Error reading PCOM file {file_path}: {e}")
        return None

def _search_patterns(text, patterns):
    """Search text using multiple patterns"""
    if isinstance(patterns, str):
        patterns = [patterns]
    
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            print(f"Pattern matched: {pattern} -> {match.group(1)}")
            return match.group(1)
    
    print(f"No pattern matched from: {patterns}")
    return None

# ============================================================
# ENHANCED MACHINE LOG PARSING (Based on Airtel code)
# ============================================================

def extract_value_enhanced(line, command):
    """
    Enhanced value extraction - similar to Airtel approach
    """
    line = line.upper().strip()
    if command in line:
        parts = line.split(command)
        if len(parts) > 1:
            value = parts[1].split('SW9000')[0] if 'SW9000' in parts[1] else parts[1]
            # Clean up - allow hex characters
            value = re.sub(r'[^0-9A-F]', '', value)
            return value
    return None

def parse_machine_log_enhanced(filepath):
    """
    Enhanced machine log parsing using Airtel's approach
    """
    print("="*80)
    print("🚀 ENHANCED MACHINE LOG PARSING")
    print("="*80)
    
    try:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()
        
        extracted = {}
        
        for line_num, line in enumerate(lines, 1):
            line_upper = line.upper().strip()
            
            # ========== ICCID (2FE2) ==========
            if '2FE2' in line_upper and ('00A40000022FE2' in line_upper or 'SELECT 2FE2' in line_upper):
                # Look ahead for ICCID data
                for i in range(min(5, len(lines) - line_num)):
                    data_line = lines[line_num + i].upper().strip()
                    if '00D600000A' in data_line:
                        val = extract_value_enhanced(data_line, '00D600000A')
                        if val and len(val) >= 20:
                            extracted['ICCID_CARD (2FE2)'] = val[:20]
                            print(f"✅ Line {line_num}: ICCID parsed: {val[:20]}")
                        break
            
            # ========== IMSI (6F07) ==========
            elif '6F07' in line_upper and ('00A40000026F07' in line_upper or 'SELECT 6F07' in line_upper):
                for i in range(min(5, len(lines) - line_num)):
                    data_line = lines[line_num + i].upper().strip()
                    if '00D6000009' in data_line:
                        val = extract_value_enhanced(data_line, '00D6000009')
                        if val and len(val) >= 18:
                            # Convert hex to decimal digits (like Airtel)
                            decimal_digits = []
                            for j in range(0, len(val), 2):
                                if j + 2 <= len(val):
                                    hex_byte = val[j:j+2]
                                    try:
                                        byte_value = int(hex_byte, 16)
                                        high_digit = (byte_value >> 4) & 0x0F
                                        low_digit = byte_value & 0x0F
                                        
                                        if 0 <= high_digit <= 9:
                                            decimal_digits.append(str(high_digit))
                                        if 0 <= low_digit <= 9:
                                            decimal_digits.append(str(low_digit))
                                    except:
                                        continue
                            
                            if decimal_digits and len(decimal_digits) >= 18:
                                imsi_18_digit = ''.join(decimal_digits)[:18]
                                extracted['HOME_IMSI (6F07)'] = imsi_18_digit
                                print(f"✅ Line {line_num}: HOME_IMSI parsed: {imsi_18_digit}")
                            elif decimal_digits:
                                imsi = ''.join(decimal_digits)
                                extracted['HOME_IMSI (6F07)'] = imsi
                                print(f"✅ Line {line_num}: HOME_IMSI parsed: {imsi}")
                        break
            
            # ========== PSK/DEK1 (6F2B) ==========
            elif '6F2B' in line_upper and ('00A40000026F2B' in line_upper or 'SELECT 6F2B' in line_upper):
                for i in range(min(5, len(lines) - line_num)):
                    data_line = lines[line_num + i].upper().strip()
                    if '00D600002AFE85410110' in data_line and 'FE80410210' in data_line:
                        parts = data_line.split('00D600002AFE85410110')
                        if len(parts) > 1:
                            tail = parts[1]
                            if 'FE80410210' in tail:
                                psk = tail.split('FE80410210')[0][:32]
                                dek1 = tail.split('FE80410210')[1][:32]
                                extracted['PSK (6F2B)'] = psk
                                extracted['DEK1 (6F2B)'] = dek1
                                print(f"✅ Line {line_num}: PSK parsed: {psk[:16]}...")
                                print(f"✅ Line {line_num}: DEK1 parsed: {dek1[:16]}...")
                        break
            
            # ========== DPUK1 (6F01) ==========
            elif '6F01' in line_upper and ('00A40000026F01' in line_upper or 'SELECT 6F01' in line_upper):
                for i in range(min(5, len(lines) - line_num)):
                    data_line = lines[line_num + i].upper().strip()
                    if '00D6000015F00A0A' in data_line:
                        val = extract_value_enhanced(data_line, '00D6000015F00A0A')
                        if val and len(val) >= 16:
                            if 'FFFFFFFF0A0A' in val:
                                parts = val.split('FFFFFFFF0A0A')
                                if len(parts) > 1 and len(parts[1]) >= 16:
                                    extracted['DPUK1_CARD (6F01)'] = parts[1][:16]
                                    print(f"✅ Line {line_num}: DPUK1 parsed: {parts[1][:16]}")
                            else:
                                extracted['DPUK1_CARD (6F01)'] = val[:16]
                                print(f"✅ Line {line_num}: DPUK1 parsed: {val[:16]}")
                        break
            
            # ========== DPUK2 (6F81) ==========
            elif '6F81' in line_upper and ('00A40000026F81' in line_upper or 'SELECT 6F81' in line_upper):
                for i in range(min(5, len(lines) - line_num)):
                    data_line = lines[line_num + i].upper().strip()
                    if '00D6000015E00A0A' in data_line:
                        val = extract_value_enhanced(data_line, '00D6000015E00A0A')
                        if val and len(val) >= 16:
                            if 'FFFFFFFF0A0A' in val:
                                parts = val.split('FFFFFFFF0A0A')
                                if len(parts) > 1 and len(parts[1]) >= 16:
                                    extracted['DPUK2_CARD (6F81)'] = parts[1][:16]
                                    print(f"✅ Line {line_num}: DPUK2 parsed: {parts[1][:16]}")
                            else:
                                extracted['DPUK2_CARD (6F81)'] = val[:16]
                                print(f"✅ Line {line_num}: DPUK2 parsed: {val[:16]}")
                        break
            
            # ========== ADM (6F0A) ==========
            elif '6F0A' in line_upper and ('00A40000026F0A' in line_upper or 'SELECT 6F0A' in line_upper):
                for i in range(min(5, len(lines) - line_num)):
                    data_line = lines[line_num + i].upper().strip()
                    if '00D600000B800A0A' in data_line:
                        val = extract_value_enhanced(data_line, '00D600000B800A0A')
                        if val and len(val) >= 16:
                            extracted['ADM (6F0A)'] = val[:16]
                            print(f"✅ Line {line_num}: ADM parsed: {val[:16]}")
                        break
            
            # ========== HOME_ACC (6F78) ==========
            elif '6F78' in line_upper and ('00A40000026F78' in line_upper or 'SELECT 6F78' in line_upper):
                for i in range(min(5, len(lines) - line_num)):
                    data_line = lines[line_num + i].upper().strip()
                    if '00D6000002' in data_line:
                        val = extract_value_enhanced(data_line, '00D6000002')
                        if val and len(val) >= 4:
                            extracted['HOME_ACC (6F78)'] = val[:4]
                            print(f"✅ Line {line_num}: HOME_ACC parsed: {val[:4]}")
                        break
            
            # ========== GLOBAL_IMSI (3031) ==========
            elif '3031' in line_upper and ('00A40000023031' in line_upper or 'SELECT 3031' in line_upper):
                for i in range(min(5, len(lines) - line_num)):
                    data_line = lines[line_num + i].upper().strip()
                    if '00D6000009' in data_line:
                        val = extract_value_enhanced(data_line, '00D6000009')
                        if val and len(val) >= 18:
                            # Convert hex to decimal digits
                            decimal_digits = []
                            for j in range(0, len(val), 2):
                                if j + 2 <= len(val):
                                    hex_byte = val[j:j+2]
                                    try:
                                        byte_value = int(hex_byte, 16)
                                        high_digit = (byte_value >> 4) & 0x0F
                                        low_digit = byte_value & 0x0F
                                        
                                        if 0 <= high_digit <= 9:
                                            decimal_digits.append(str(high_digit))
                                        if 0 <= low_digit <= 9:
                                            decimal_digits.append(str(low_digit))
                                    except:
                                        continue
                            
                            if decimal_digits and len(decimal_digits) >= 18:
                                imsi_18_digit = ''.join(decimal_digits)[:18]
                                extracted['GLOBAL_IMSI (3031)'] = imsi_18_digit
                                print(f"✅ Line {line_num}: GLOBAL_IMSI parsed: {imsi_18_digit}")
                        break
            
            # ========== GLOBAL_ACC (3037) ==========
            elif '3037' in line_upper and ('00A40000023037' in line_upper or 'SELECT 3037' in line_upper):
                for i in range(min(5, len(lines) - line_num)):
                    data_line = lines[line_num + i].upper().strip()
                    if '00D60000120000000300000002FFFFFFFF' in data_line:
                        val = extract_value_enhanced(data_line, '00D60000120000000300000002FFFFFFFF')
                        if val and len(val) >= 8:
                            extracted['GLOBAL_ACC (3037)'] = val[:4]
                            extracted['HOME_ACC (3037)'] = val[4:8]
                            print(f"✅ Line {line_num}: GLOBAL_ACC parsed: {val[:4]}")
                            print(f"✅ Line {line_num}: HOME_ACC parsed: {val[4:8]}")
                        break
            
            # ========== KIC/KID Keys (6F22) ==========
            elif '6F22' in line_upper and ('00A40000026F22' in line_upper or 'SELECT 6F22' in line_upper):
                kic_kid_prefixes = {
                    'KIC1 (6F22)': '00DC01041BFE0150',
                    'KID1 (6F22)': '00DC02041BFE0151',
                    'KIK1 (6F22)': '00DC03041BFE0152',
                    'KIC2 (6F22)': '00DC04041BFE0250',
                    'KID2 (6F22)': '00DC05041BFE0251',
                    'KIK2 (6F22)': '00DC06041BFE0252',
                }
                
                for prefix_key, prefix in kic_kid_prefixes.items():
                    for i in range(min(10, len(lines) - line_num)):
                        data_line = lines[line_num + i].upper().strip()
                        if prefix in data_line:
                            val = extract_value_enhanced(data_line, prefix)
                            if val:
                                val = val.replace('FFFFFFFFFFFFFFFF', '')
                                if len(val) >= 32:
                                    extracted[prefix_key] = val[:32]
                                    print(f"✅ Line {line_num+i+1}: {prefix_key} parsed: {val[:16]}...")
                            break
            
            # ========== ASCII IMSI (6F02) ==========
            elif '6F02' in line_upper and ('00A40000026F02' in line_upper or 'SELECT 6F02' in line_upper):
                for i in range(min(5, len(lines) - line_num)):
                    data_line = lines[line_num + i].upper().strip()
                    if '00D600005F8031' in data_line:
                        parts = data_line.split('00D600005F8031')
                        if len(parts) > 1:
                            val = parts[1]
                            extracted['ASCII_IMSI (6F02)'] = val[:30]
                            print(f"✅ Line {line_num}: ASCII_IMSI (6F02) parsed: {val[:30]}")
                        break
            
            # ========== ASCII IMSI (6F04) ==========
            elif '6F04' in line_upper and ('00A40000026F04' in line_upper or 'SELECT 6F04' in line_upper):
                for i in range(min(5, len(lines) - line_num)):
                    data_line = lines[line_num + i].upper().strip()
                    if '00DC01047880357369703A' in data_line:
                        parts = data_line.split('00DC01047880357369703A')
                        if len(parts) > 1:
                            val = parts[1]
                            extracted['ASCII_IMSI (6F04)'] = val[:30]
                            print(f"✅ Line {line_num}: ASCII_IMSI (6F04) parsed: {val[:30]}")
                        break
        
        print(f"\n✅ Extracted {len(extracted)} fields from Machine Log")
        for field, value in extracted.items():
            print(f"   - {field}: {value}")
        
        return extracted
        
    except Exception as e:
        print(f"❌ Machine Log parsing error: {str(e)}")
        traceback.print_exc()
        return {}

def parse_machine_log_robust(filepath):
    """
    Robust machine log parsing with multiple strategies
    """
    print("\n" + "="*80)
    print("ROBUST MACHINE LOG PARSING")
    print("="*80)
    
    # Strategy 1: Enhanced parsing (Airtel style)
    results1 = parse_machine_log_enhanced(filepath)
    
    # Strategy 2: Pattern-based search (fallback)
    results2 = {}
    try:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
        
        # Pattern search for specific fields
        patterns = [
            ('ICCID_CARD (2FE2)', r'00D600000A([0-9A-F]{20})'),
            ('HOME_IMSI (6F07)', r'00D6000009([0-9A-F]{18})'),
            ('PSK (6F2B)', r'FE85410110([0-9A-F]{32})'),
            ('DEK1 (6F2B)', r'FE80410210([0-9A-F]{32})'),
            ('DPUK1_CARD (6F01)', r'F00A0A([0-9A-F]{16})'),
            ('DPUK2_CARD (6F81)', r'E00A0A([0-9A-F]{16})'),
            ('ADM (6F0A)', r'800A0A([0-9A-F]{16})'),
            ('HOME_ACC (6F78)', r'00D6000002([0-9A-F]{4})'),
            ('GLOBAL_IMSI (3031)', r'00A40000023031[^\n]*?\n[^\n]*?00D6000009([0-9A-F]{18})'),
        ]
        
        for field, pattern in patterns:
            matches = re.findall(pattern, content, re.IGNORECASE)
            if matches:
                results2[field] = matches[0]
                print(f"✅ Pattern found {field}: {matches[0]}")
    
    except Exception as e:
        print(f"Pattern search error: {e}")
    
    # Merge results (prefer results1, fallback to results2)
    all_results = {}
    all_results.update(results2)  # Pattern results first
    all_results.update(results1)  # Enhanced results override
    
    print(f"\n📊 Total extracted: {len(all_results)} fields")
    return all_results

class ValidationEngine:
    def __init__(self):
        self.results = {}

def main(profile_type, filepath, pcom_path, cnum_path, scm_path, sim_oda_path, image_paths=None):
    
    print("Running with:", profile_type, filepath)

    # Initialize error collection
    validation_errors = []

    # Fields to skip per profile
    skip_map = {
        "MOB": [],
        "WBIOT": ["GLOBAL_IMSI (3031)", "GLOBAL_ACC (3037)", "HOME_ACC (3037)"],
        "NBIOT": ["GLOBAL_IMSI (3031)", "GLOBAL_ACC (3037)", "HOME_ACC (3037)", 
                  "ASCII_IMSI (6F02)", "ASCII_IMSI (6F04)"]
    }

    skip_fields = set(skip_map.get(profile_type, []))

    # Field definitions
    key_to_header = {
        "PSK (6F2B)": "PSK (6F2B)",
        "DEK1 (6F2B)": "DEK1 (6F2B)",
        "GLOBAL_IMSI (3031)": "GLOBAL_IMSI (3031)",
        "HOME_IMSI (6F07)": "HOME_IMSI (6F07)",
        "GLOBAL_ACC (3037)": "GLOBAL_ACC (3037)",
        "HOME_ACC (6F78)": "HOME_ACC (6F78)",
        "DPUK1_CARD (6F01)": "DPUK1_CARD (6F01)",
        "DPUK2_CARD (6F81)": "DPUK2_CARD (6F81)",
        "ADM (6F0A)": "ADM (6F0A)",
        "ICCID_CARD (2FE2)": "ICCID_CARD (2FE2)",
        "KIC1 (6F22)": "KIC1 (6F22)",
        "KID1 (6F22)": "KID1 (6F22)",
        "KIC2 (6F22)": "KIC2 (6F22)",
        "KID2 (6F22)": "KID2 (6F22)",
        "ASCII_IMSI (6F02)": "ASCII_IMSI (6F02)",
        "ASCII_IMSI (6F04)": "ASCII_IMSI (6F04)",
        "HOME_ACC (3037)": "HOME_ACC (3037)",
    }

    # Enhanced PCOM configuration with multiple search patterns
    if profile_type == 'NBIOT':
        pcom_config = {
            "HOME_IMSI (6F07)": [
                r"\.DEFINE\s+%IMSI\s+([0-9]+)",
                r"\.DEFINE\s+%HOME_IMSI\s+([0-9]+)",
                r"IMSI\s*=\s*([0-9]+)"
            ],
            "HOME_ACC (6F78)": [
                r"\.DEFINE\s+%ACC\s+([0-9]+)", 
                r"\.DEFINE\s+%HOME_ACC\s+([0-9]+)",
                r"ACC\s*=\s*([0-9]+)"
            ],
            "DPUK1_CARD (6F01)": [
                r"\.DEFINE\s+%PUK1\s+([0-9]+)",
                r"\.DEFINE\s+%DPUK1_CARD\s+([0-9]+)",
                r"PUK1\s*=\s*([0-9]+)"
            ],
            "DPUK2_CARD (6F81)": [
                r"\.DEFINE\s+%PUK2\s+([0-9]+)",
                r"\.DEFINE\s+%DPUK2_CARD\s+([0-9]+)",
                r"PUK2\s*=\s*([0-9]+)"
            ],
            "ADM (6F0A)": [
                r"\.DEFINE\s+%ISC1\s+([0-9A-Fa-f]+)",
                r"\.DEFINE\s+%ADM1_CARD\s+([0-9A-Fa-f]+)",
                r"ADM\s*=\s*([0-9A-Fa-f]+)"
            ],
            "ICCID_CARD (2FE2)": [
                r"\.DEFINE\s+%ICCID\s+([0-9]+)",
                r"\.DEFINE\s+%ICCID_CARD\s+([0-9]+)",
                r"ICCID\s*=\s*([0-9]+)"
            ]
        }
    else: 
        pcom_config = {
            "GLOBAL_IMSI (3031)": [
                r"\.DEFINE\s+%HOME_IMSI\s+([0-9]+)",
                r"IMSI\s*=\s*([0-9]+)"
            ],
            "HOME_IMSI (6F07)": [
                r"\.DEFINE\s+%HOME_IMSI\s+([0-9]+)",
                r"IMSI\s*=\s*([0-9]+)"
            ],
            "GLOBAL_ACC (3037)": [
                r"\.DEFINE\s+%HOME_ACC\s+([0-9]+)",
                r"ACC\s*=\s*([0-9]+)"
            ],
            "HOME_ACC (6F78)": [
                r"\.DEFINE\s+%HOME_ACC\s+([0-9]+)",
                r"ACC\s*=\s*([0-9]+)"
            ],
            "DPUK1_CARD (6F01)": [
                r"\.DEFINE\s+%PUK1\s+([0-9]+)",
                r"\.DEFINE\s+%DPUK1_CARD\s+([0-9]+)",
                r"PUK1\s*=\s*([0-9]+)"
            ],
            "DPUK2_CARD (6F81)": [
                r"\.DEFINE\s+%PUK2\s+([0-9]+)",
                r"\.DEFINE\s+%DPUK2_CARD\s+([0-9]+)",
                r"PUK2\s*=\s*([0-9]+)"
            ],
            "ADM (6F0A)": [
                r"\.DEFINE\s+%ISC1\s+([0-9A-Fa-f]+)",
                r"\.DEFINE\s+%ADM1_CARD\s+([0-9A-Fa-f]+)",
                r"ADM\s*=\s*([0-9A-Fa-f]+)"
            ],
            "ICCID_CARD (2FE2)": [
                r"\.DEFINE\s+%ICCID\s+([0-9]+)",
                r"\.DEFINE\s+%ICCID_CARD\s+([0-9]+)",
                r"ICCID\s*=\s*([0-9]+)"
            ],
            "ASCII_IMSI (6F02)": [
                r"\.DEFINE\s+%ASCII_IMSI\s+([0-9]+)",
                r"ASCII_IMSI\s*=\s*([0-9]+)"
            ],
            "ASCII_IMSI (6F04)": [
                r"\.DEFINE\s+%ASCII_IMSI\s+([0-9]+)",
                r"ASCII_IMSI\s*=\s*([0-9]+)"
            ],
            "HOME_ACC (3037)": [
                r"\.DEFINE\s+%HOME_ACC\s+([0-9]+)",
                r"ACC\s*=\s*([0-9]+)"
            ]
        }
    
    cnum_config = {
        "GLOBAL_IMSI (3031)": (16, 2, True),  # Special logic, only first value
        "HOME_IMSI (6F07)": (16, 2, True),    # Special logic, only first value
        "ASCII_IMSI (6F02)": (16, 2, True),   # Special logic, only first value
        "ASCII_IMSI (6F04)": (16, 2, True),   # Special logic, only first value
        "ICCID_CARD (2FE2)": (16, 4, False),  # Standard logic
        "DPUK1_CARD (6F01)": (16, 6, False),  # Standard logic
        "DPUK2_CARD (6F81)": (16, 8, False)   # Standard logic
    }

    scm_config = {
        "GLOBAL_IMSI (3031)": (2, 3),   # Column 4 (0-indexed)
        "HOME_IMSI (6F07)": (2, 3),
        "ICCID_CARD (2FE2)": (2, 2),    # Column 3
        "ASCII_IMSI (6F02)": (2, 3),
        "ASCII_IMSI (6F04)": (2, 3)
    }
    
    if profile_type == 'NBIOT':
        sim_oda_config = {
            "PSK (6F2B)": (126, r"SecurityKey\([^,]+, [^,]+, PskTls, ([^,]+),"),
            "DEK1 (6F2B)": (127, r"SecurityKey\([^,]+, [^,]+, Management, ([^,]+),"),
            "GLOBAL_IMSI (3031)": (130, r"Imsi\((\w+)\)"),
            "HOME_IMSI (6F07)": (130, r"Imsi\((\w+)\)"),
            "ICCID_CARD (2FE2)": (121, r"Iccid\(([^,]+),.*\)"),
            "KIC1 (6F22)": (122, r"SecurityKey\(.*, Encryption, (\w+)\)"),
            "KID1 (6F22)": (123, r"SecurityKey\(.*, Authentication, (\w+)\)"),
            "KIC2 (6F22)": (124, r"SecurityKey\(.*, Encryption, (\w+)\)"),
            "KID2 (6F22)": (125, r"SecurityKey\(.*, Authentication, (\w+)\)"),
            "ASCII_IMSI (6F02)": (130, r"Imsi\((\w+)\)"),
            "ASCII_IMSI (6F04)": (130, r"Imsi\((\w+)\)"),
        }
    else:
        sim_oda_config = {
            "PSK (6F2B)": (351, r"SecurityKey\([^,]+, [^,]+, PskTls, ([^,]+),"),
            "DEK1 (6F2B)": (352, r"SecurityKey\([^,]+, [^,]+, Management, ([^,]+),"),
            "GLOBAL_IMSI (3031)": (355, r"Imsi\((\w+)\)"),
            "HOME_IMSI (6F07)": (355, r"Imsi\((\w+)\)"),
            "ICCID_CARD (2FE2)": (346, r"Iccid\(([^,]+),.*\)"),
            "KIC1 (6F22)": (347, r"SecurityKey\(.*, Encryption, (\w+)\)"),
            "KID1 (6F22)": (348, r"SecurityKey\(.*, Authentication, (\w+)\)"),
            "KIC2 (6F22)": (349, r"SecurityKey\(.*, Encryption, (\w+)\)"),
            "KID2 (6F22)": (350, r"SecurityKey\(.*, Authentication, (\w+)\)"),
            "ASCII_IMSI (6F02)": (355, r"Imsi\((\w+)\)"),
            "ASCII_IMSI (6F04)": (355, r"Imsi\((\w+)\)")
        }

    # Extract values from all files
    file_values = {k: {"PCOM": None, "CNUM": None, "SCM": None, "SIM_ODA": None} 
                for k in key_to_header if k not in skip_fields}
    
    # ============================================================
    # ENHANCED PCOM EXTRACTION
    # ============================================================
    print("\n" + "="*80)
    print("EXTRACTING FROM PCOM FILE")
    print("="*80)
    
    for key, patterns in pcom_config.items():
        if key in skip_fields:
            continue
        
        print(f"\n🔍 Searching for {key} in PCOM...")
        extracted_value = extract_from_pcom_enhanced(pcom_path, patterns)
        file_values[key]["PCOM"] = extracted_value
        
        if extracted_value is None:
            error_msg = f"[{key}] PCOM extraction failed"
            print(f"❌ {error_msg}")
            validation_errors.append(error_msg)
        else:
            print(f"✅ [{key}] PCOM extraction successful: {extracted_value}")

    # ============================================================
    # ENHANCED MACHINE LOG PARSING
    # ============================================================
    print("\n" + "="*80)
    print("EXTRACTING FROM MACHINE LOG")
    print("="*80)
    
    results = parse_machine_log_robust(filepath)
    
    # Initialize all fields
    for key in key_to_header.keys():
        if key not in results:
            results[key] = "N/A"
    
    # ============================================================
    # EXTRACT FROM OTHER FILES
    # ============================================================
    
    # First handle KIC/KID keys using ordered match
    kic_matches = extract_multiple_keys(sim_oda_path, r"SecurityKey\(.*, Encryption, (\w+)\)")
    kid_matches = extract_multiple_keys(sim_oda_path, r"SecurityKey\(.*, Authentication, (\w+)\)")

    if "KIC1 (6F22)" in file_values:
        file_values["KIC1 (6F22)"]["SIM_ODA"] = kic_matches[0] if len(kic_matches) > 0 else None
    if "KIC2 (6F22)" in file_values:
        file_values["KIC2 (6F22)"]["SIM_ODA"] = kic_matches[1] if len(kic_matches) > 1 else None
    if "KID1 (6F22)" in file_values:
        file_values["KID1 (6F22)"]["SIM_ODA"] = kid_matches[0] if len(kid_matches) > 0 else None
    if "KID2 (6F22)" in file_values:
        file_values["KID2 (6F22)"]["SIM_ODA"] = kid_matches[1] if len(kid_matches) > 1 else None

    # Extract from CNUM
    for key, (line_num, col_idx, special_logic) in cnum_config.items():
        if key in skip_fields:
            continue
        if key in file_values:
            file_values[key]["CNUM"] = extract_from_cnum(cnum_path, line_num, col_idx, special_logic)

    # Extract from SCM
    for key, (line_num, col_idx) in scm_config.items():
        if key in skip_fields:
            continue
        if key in file_values:
            file_values[key]["SCM"] = extract_from_scm(scm_path, line_num, col_idx)

    # Extract from SIM ODA
    for key, (line_num, pattern) in sim_oda_config.items():
        if key not in file_values or key in ["KIC1 (6F22)", "KIC2 (6F22)", "KID1 (6F22)", "KID2 (6F22)"]:
            continue
        file_values[key]["SIM_ODA"] = extract_from_sim_oda(
            sim_oda_path, line_num, pattern, search_range=2, fallback=True
        )

    # ============================================================
    # VALIDATION RULES
    # ============================================================
    rules = {
        "PSK (6F2B)": ("NR", "NR", "NR", "from_value"),
        "DEK1 (6F2B)": ("NR", "NR", "NR", "from_value"),
        "GLOBAL_IMSI (3031)": ("from_value", "from_value", "from_value", "from_value"),
        "HOME_IMSI (6F07)": ("from_value", "from_value", "from_value", "from_value"),
        "GLOBAL_ACC (3037)": ("from_value", "NR", "NR", "NR"),
        "HOME_ACC (6F78)": ("from_value", "NR", "NR", "NR"),
        "DPUK1_CARD (6F01)": ("from_value", "from_value", "NR", "NR"),
        "DPUK2_CARD (6F81)": ("from_value", "from_value", "NR", "NR"),
        "ADM (6F0A)": ("from_value", "NR", "NR", "NR"),
        "ICCID_CARD (2FE2)": ("from_value", "from_value", "from_value", "from_value"),
        "KIC1 (6F22)": ("NR", "NR", "NR", "from_value"),
        "KID1 (6F22)": ("NR", "NR", "NR", "from_value"),
        "KIC2 (6F22)": ("NR", "NR", "NR", "from_value"),
        "KID2 (6F22)": ("NR", "NR", "NR", "from_value"),
        "ASCII_IMSI (6F02)": ("from_value", "from_value", "from_value", "from_value"),
        "ASCII_IMSI (6F04)": ("from_value", "from_value", "from_value", "from_value"),
        "HOME_ACC (3037)": ("from_value", "NR", "NR", "NR"),
    }

    # Excel Report Setup
    wb = Workbook()
    ws = wb.active
    ws.title = "Validation Report"

    if image_paths:
        # Set starting row based on profile_type
        if profile_type == "NBIOT":
            start_row = 30
        elif profile_type == "WBIOT":
            start_row = 32
        else:
            start_row = 30  # Default start row

        start_col_letter = 'D'  # Column D

        for i, image_path in enumerate(image_paths):
            if image_path:  # If a valid image path is provided
                row = start_row + (i * 10)  # Increase row by 10 for each image
                cell = f'{start_col_letter}{row}'  # e.g., D32, D42, D52, etc.
                insert_image(ws, image_path, cell)  # Insert image at the specified cell

    # Styles
    styles = setup_excel_styles()

    # Headers
    headers = ["Field", "Machine Log", "PCOM", "CNUM", "SCM", "SIM ODA", 
               "PCOM Status", "CNUM Status", "SCM Status", "SIM_ODA Status", "Validation Status"]
    
    # Setup Excel headers and metadata
    setup_excel_headers(ws, styles)

    header_row = 16  # Row 18 will now contain your table headers
           
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.fill = styles['dark_blue_fill']
        cell.font = styles['header_font']
        cell.border = styles['thick_border']
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Iterate over each key
    row = header_row + 1 

    mismatch_logs = []  

    for key in key_to_header.keys():
        if key in skip_fields:
            continue

        ml_val = results.get(key, "N/A")
        rule = rules.get(key, ("NR", "NR", "NR", "NR"))

        pcom_raw = file_values.get(key, {}).get("PCOM")
        cnum_raw = file_values.get(key, {}).get("CNUM")
        scm_raw = file_values.get(key, {}).get("SCM")
        sim_oda_raw = file_values.get(key, {}).get("SIM_ODA")

        pcom_disp = pcom_raw if rule[0] == "from_value" and pcom_raw else "Missing" if rule[0] == "from_value" else "NR"
        cnum_disp = cnum_raw if rule[1] == "from_value" and cnum_raw else "Missing" if rule[1] == "from_value" else "NR"
        scm_disp = scm_raw if rule[2] == "from_value" and scm_raw else "Missing" if rule[2] == "from_value" else "NR"
        sim_oda_disp = sim_oda_raw if rule[3] == "from_value" and sim_oda_raw else "Missing" if rule[3] == "from_value" else "NR"

        status = {"PCOM": "NR", "CNUM": "NR", "SCM": "NR", "SIM_ODA": "NR"}
        overall_valid = True

        def is_ascii_imsi_key(k): return "ASCII_IMSI" in k
        def is_imsi_key(k): return "IMSI" in k and "ASCII_IMSI" not in k
        def is_iccid_key(k): return k == "ICCID_CARD (2FE2)"
        def is_dpuk_key(k): return k in ["DPUK1_CARD (6F01)", "DPUK2_CARD (6F81)"]

        if rule[0] == "from_value":
            if pcom_raw:
                if len(pcom_raw) != len(ml_val):
                    error_msg = f"[{key}] PCOM length mismatch: ML length={len(ml_val)}, PCOM length={len(pcom_raw)}"
                    mismatch_logs.append(
                        f"[{key}] PCOM length mismatch: ML length={len(ml_val)}, PCOM length={len(pcom_raw)}, "
                        f"ML value={ml_val}, PCOM value={pcom_raw}"
                    )
                    validation_errors.append(error_msg)
                    status["PCOM"] = "❌ Fail"
                    overall_valid = False
                else:
                    if is_iccid_key(key):
                        norm_ml, norm = normalize_iccid(ml_val), normalize_iccid(pcom_raw)
                    elif is_ascii_imsi_key(key):
                        norm_ml, norm = normalize_ascii_imsi(ml_val), normalize_ascii_imsi(pcom_raw)
                    elif is_imsi_key(key):
                        norm_ml, norm = normalize_imsi(ml_val), normalize_imsi(pcom_raw)
                    else:
                        norm_ml, norm = ml_val, pcom_raw
                    if norm_ml == norm:
                        status["PCOM"] = "✅ Pass"
                    else:
                        status["PCOM"] = "❌ Fail"
                        error_msg = f"[{key}] PCOM mismatch: ML={norm_ml}, PCOM={norm}"
                        mismatch_logs.append(error_msg)
                        validation_errors.append(error_msg)
                        overall_valid = False
            else:
                status["PCOM"] = "❌ Fail"
                error_msg = f"[{key}] PCOM Missing"
                mismatch_logs.append(error_msg)
                validation_errors.append(error_msg)
                overall_valid = False

        # DPUK special handling
        if is_dpuk_key(key):
            try:
                decoded_ml = bytes.fromhex(ml_val).decode('ascii')
            except ValueError:
                decoded_ml = ""

            norm_cnum = re.sub(r'\D', '', cnum_raw or "")
            norm_pcom = re.sub(r'\D', '', pcom_raw or "")

            status = {
                "PCOM": "✅ Pass" if ml_val == norm_pcom else "❌ Fail" if pcom_raw else "Missing",
                "CNUM": "✅ Pass" if decoded_ml == norm_cnum else "❌ Fail" if cnum_raw else "Missing",
                "SCM": "NR",
                "SIM_ODA": "NR"
            }

            if status["PCOM"] != "✅ Pass":
                error_msg = f"[{key}] DPUK PCOM mismatch: ML={ml_val}, PCOM={norm_pcom}"
                mismatch_logs.append(error_msg)
                validation_errors.append(error_msg)
            if status["CNUM"] != "✅ Pass":
                error_msg = f"[{key}] DPUK CNUM mismatch: Decoded ML={decoded_ml}, CNUM={norm_cnum}"
                mismatch_logs.append(error_msg)
                validation_errors.append(error_msg)

            overall_valid = all(v == "✅ Pass" or v == "NR" for v in status.values())

            # Display values
            pcom_disp = pcom_raw or "Missing"
            cnum_disp = cnum_raw or "Missing"
            scm_disp = sim_oda_disp = "NR"
            validation_status = "✅ Pass" if overall_valid else "❌ Fail"
            data = [key, ml_val, pcom_disp, cnum_disp, scm_disp, sim_oda_disp,
                    status["PCOM"], status["CNUM"], status["SCM"], status["SIM_ODA"], validation_status]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = styles['thick_border']
                cell.number_format = '@'
                if "Pass" in str(value):
                    cell.fill = styles['green_fill']
                elif "Fail" in str(value):
                    cell.fill = styles['red_fill']
                elif str(value).strip().upper() in ["NR", "N/A", "Missing"]:
                    cell.fill = styles['yellow_fill']
            row += 1
            continue

        # CNUM validation
        if rule[1] == "from_value":
            if cnum_raw:
                if is_iccid_key(key):
                    norm_ml = normalize_iccid(ml_val)
                    norm = re.sub(r'\D', '', cnum_raw)
                elif is_ascii_imsi_key(key):
                    norm_ml, norm = normalize_ascii_imsi(ml_val), cnum_raw
                elif is_imsi_key(key):
                    norm_ml, norm = normalize_imsi(ml_val), re.sub(r'\D', '', cnum_raw)
                else:
                    norm_ml, norm = ml_val, cnum_raw
                if norm_ml == norm and len(norm_ml) == len(norm):
                    status["CNUM"] = "✅ Pass"
                else:
                    status["CNUM"] = "❌ Fail"
                    error_msg = f"[{key}] CNUM mismatch: ML={norm_ml}, CNUM={norm}"
                    mismatch_logs.append(error_msg)
                    validation_errors.append(error_msg)
                    overall_valid = False
            else:
                status["CNUM"] = "❌ Fail"
                error_msg = f"[{key}] CNUM Missing"
                mismatch_logs.append(error_msg)
                validation_errors.append(error_msg)
                overall_valid = False

        # SCM validation
        if rule[2] == "from_value":
            if scm_raw:
                if is_iccid_key(key):
                    norm_ml = normalize_iccid(ml_val)
                    norm = re.sub(r'\D', '', scm_raw)
                elif is_ascii_imsi_key(key):
                    norm_ml, norm = normalize_ascii_imsi(ml_val), scm_raw
                elif is_imsi_key(key):
                    norm_ml, norm = normalize_imsi(ml_val), re.sub(r'\D', '', scm_raw)
                else:
                    norm_ml, norm = ml_val, scm_raw
                if norm_ml == norm and len(norm_ml) == len(norm):
                    status["SCM"] = "✅ Pass"
                else:
                    status["SCM"] = "❌ Fail"
                    error_msg = f"[{key}] SCM mismatch: ML={norm_ml}, SCM={norm}"
                    mismatch_logs.append(error_msg)
                    validation_errors.append(error_msg)
                    overall_valid = False
            else:
                status["SCM"] = "❌ Fail"
                error_msg = f"[{key}] SCM Missing"
                mismatch_logs.append(error_msg)
                validation_errors.append(error_msg)
                overall_valid = False

        # SIM_ODA validation
        if rule[3] == "from_value":
            if sim_oda_raw:
                if is_iccid_key(key):
                    norm_ml = normalize_iccid(ml_val)
                    norm = re.sub(r'\D', '', sim_oda_raw)
                elif is_ascii_imsi_key(key):
                    norm_ml, norm = normalize_ascii_imsi(ml_val), sim_oda_raw
                elif is_imsi_key(key):
                    norm_ml, norm = normalize_imsi(ml_val), re.sub(r'\D', '', sim_oda_raw)
                else:
                    norm_ml, norm = ml_val, sim_oda_raw
                if norm_ml == norm and len(norm_ml) == len(norm):
                    status["SIM_ODA"] = "✅ Pass"
                else:
                    status["SIM_ODA"] = "❌ Fail"
                    error_msg = f"[{key}] SIM_ODA mismatch: ML={norm_ml}, SIM_ODA={norm}"
                    mismatch_logs.append(error_msg)
                    validation_errors.append(error_msg)
                    overall_valid = False
            else:
                status["SIM_ODA"] = "❌ Fail"
                error_msg = f"[{key}] SIM_ODA Missing"
                mismatch_logs.append(error_msg)
                validation_errors.append(error_msg)
                overall_valid = False

        validation_status = "✅ Pass" if overall_valid else "❌ Fail"

        # Write to Excel
        data = [
            key, ml_val, pcom_disp, cnum_disp, scm_disp, sim_oda_disp,
            status["PCOM"], status["CNUM"], status["SCM"], status["SIM_ODA"], validation_status
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = styles['thick_border']
            cell.number_format = '@'
            if "Pass" in str(value):
                cell.fill = styles['green_fill']
            elif "Fail" in str(value):
                cell.fill = styles['red_fill']
            elif str(value).strip().upper() in ["NR", "N/A", "Missing"]:
                cell.fill = styles['yellow_fill']

        row += 1

    # --- Begin processing ---
    image_label_map = {
        "INNER LABEL 100": image_paths[0] if len(image_paths) > 0 else None,
        "INNER LABEL 500": image_paths[1] if len(image_paths) > 1 else None,
        "OUTER LABEL 5000": image_paths[2] if len(image_paths) > 2 else None,
        "ARTWORK FRONT": image_paths[3] if len(image_paths) > 3 else None,
        "ARTWORK BACK": image_paths[4] if len(image_paths) > 4 else None,
    }

    print("Image Label Map:", image_label_map)

    label_qr_data = {}
    if profile_type == "MOB":
        for label, img_path in image_label_map.items():
            if img_path and os.path.isfile(img_path):
                print(f"--- Scanning {label} (WBIOT/NBIOT) from {img_path}")
                raw_qr_data = process_qr_code_wbiot(img_path) or {}

                # ✅ Normalize the keys immediately
                normalized_qr_data = {
                    k.upper().replace(" ", "").replace(".", "").replace("_", ""): v
                    for k, v in raw_qr_data.items()
                }
                print(f"QR parsed data for {label}:", normalized_qr_data)
                label_qr_data[label] = normalized_qr_data
            else:
                print(f"--- Skipping {label}, file missing or invalid: {img_path}")
                label_qr_data[label] = {}

        # Merge OUTER + ARTWORK for MOB if needed
        merged_outer = label_qr_data.get("OUTER LABEL 5000", {}).copy()
        for k, v in label_qr_data.get("ARTWORK FRONT", {}).items():
            merged_outer.setdefault(k, v)
        for k, v in label_qr_data.get("ARTWORK BACK", {}).items():
            merged_outer.setdefault(k, v)
        label_qr_data["OUTER LABEL 5000"] = merged_outer

        label_sections = [
            ("INNER LABEL 500", ["ICCID Start", "ICCID End", "PO", "QTY", "BATCH NO", "EAN", "CIRCLE", "PRODUCT","MSC","MSN"]),
            ("OUTER LABEL 5000", ["ICCID Start", "ICCID End", "PO", "QTY", "EAN", "CIRCLE", "BATCH NO", "PRODUCT" ,"MSN1", "MSN10", "MSC"]),
            ("ARTWORK", ["ICCID"]),
        ]

    elif profile_type in ["WBIOT", "NBIOT"]: 
        for label, img_path in image_label_map.items():
            if img_path and os.path.isfile(img_path):
                print(f"--- Scanning {label} (WBIOT/NBIOT) from {img_path}")
                qr_data = process_qr_code_wbiot(img_path) or {}

                # ✅ Normalize all keys here (so ICCID and others work in report)
                normalized_qr_data = {
                    k.upper().replace(" ", "").replace(".", "").replace("_", ""): v
                    for k, v in qr_data.items()
                }

                print(f"QR parsed data for {label}:", normalized_qr_data)
                label_qr_data[label] = normalized_qr_data

            else:
                print(f"--- Skipping {label}, file missing or invalid: {img_path}")
                label_qr_data[label] = {}

        # Merge ARTWORK FRONT and BACK
        artwork_merged = {}
        for key in ["ARTWORK FRONT", "ARTWORK BACK"]:
            for k, v in label_qr_data.get(key, {}).items():
                artwork_merged[k] = v
        label_qr_data["ARTWORK"] = artwork_merged

        # Excel section layout definition
        label_sections = [
            ("INNER LABEL 100", ["ICCID Start", "ICCID End", "PO", "BATCH NO", "QTY", "CIRCLE", "EAN", "PID", "Product", "MSN"]),
            ("INNER LABEL 500", ["ICCID Start", "ICCID End", "PO", "QTY", "BATCH NO", "EAN", "CIRCLE", "PRODUCT","MSC","MSN"]),
            ("OUTER LABEL 5000", ["ICCID Start", "ICCID End", "PO", "QTY", "EAN", "CIRCLE", "BATCH NO", "PRODUCT" ,"MSN1", "MSN10", "MSC"]),
            ("ARTWORK", ["ICCID"]),
        ]

    else:
        print("❌ Unsupported profile type")
        return None, validation_errors

    # Styles
    label_fill = PatternFill(start_color='002060', end_color='002060', fill_type='solid')
    label_font = Font(color='FFFFFF', bold=True, size=12)
    label_align = Alignment(horizontal='left', vertical='center')
    bold_border = Border(
        left=Side(style='thick'), right=Side(style='thick'),
        top=Side(style='thick'), bottom=Side(style='thick')
    )

    # Start row (after main table)
    label_start_row = row + 1  # 'row' must already be defined

    # Write sections
    for idx, (section_title, fields) in enumerate(label_sections):
        # Section header
        for col in range(1, 4):
            cell = ws.cell(row=label_start_row, column=col)
            cell.fill = label_fill
            cell.font = label_font
            cell.alignment = label_align
            cell.border = bold_border

        ws.cell(row=label_start_row, column=1, value=section_title)
        ws.cell(row=label_start_row, column=2, value="Values")
        ws.cell(row=label_start_row, column=3, value="Status")

        # Get normalized data
        section_data_raw = label_qr_data.get(section_title, {})
        section_data = {
            k.upper().replace(" ", "").replace(".", ""): v
            for k, v in section_data_raw.items()
        }

        # Section fields
        for field in fields:
            label_start_row += 1
            ws.cell(row=label_start_row, column=1, value=field).border = bold_border

            normalized_field = field.upper().replace(" ", "").replace(".", "")
            value = section_data.get(normalized_field, "")
            ws.cell(row=label_start_row, column=2, value=value).border = bold_border
            ws.cell(row=label_start_row, column=3).border = bold_border

        # Add space after section (except last one)
        if idx < len(label_sections) - 1:
            label_start_row += 2

    # --- Auto-Fit Columns Based on Actual Content ---

    for col_idx, _ in enumerate(headers, 1):  # Start at 1
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, label_start_row + 1)
        )
        ws.column_dimensions[col_letter].width = max_len + 8

    # --- Save Report ---

    # After saving report
    report_path = save_report(wb, filepath, pcom_path)
    print(f"\n" + "="*80)
    print(f"REPORT GENERATION COMPLETE")
    print(f"Report saved at: {report_path}")
    print(f"Total validation errors: {len(validation_errors)}")
    print("="*80)
    
    return report_path, validation_errors